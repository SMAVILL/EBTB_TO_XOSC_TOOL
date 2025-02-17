import tkinter as tk
from pathlib import Path
from tkinter import filedialog
import webbrowser
import os
from datetime import datetime
import getpass
from collections import defaultdict
import re

selected_path = ""


# Function to create the GUI and return the selected path
def ebtb_GUI():
    selected_path = ""  # Initialize variable to store the selected path

    def get_swc(data_model):
        nonlocal selected_path
        selected_path = data_model
        print(f"Selected path stored: {selected_path}")

    # GUI Code
    def select_folder():
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            data_model_entry.delete(0, tk.END)  # Clear any existing text
            data_model_entry.insert(0, folder_selected)  # Insert selected path

    def run_get_swc():
        data_model = data_model_entry.get()
        if data_model:
            get_swc(data_model)
            root.quit()  # End the Tkinter main loop to close the window

    # Main GUI setup
    root = tk.Tk()
    root.title("EBTB_to_XOSC_Tool_ver1.1")

    # Input label and entry field
    tk.Label(root, text="Select EBTB_Folder:").grid(row=0, column=0, padx=10, pady=10)
    data_model_entry = tk.Entry(root, width=50)
    data_model_entry.grid(row=0, column=1, padx=10, pady=10)

    # Button to open folder dialog
    browse_button = tk.Button(root, text="Browse", command=select_folder)
    browse_button.grid(row=0, column=2, padx=10, pady=10)

    # Button to run the function and store the path, then end the program
    run_button = tk.Button(root, text="Convert", command=run_get_swc)
    run_button.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

    # Start the GUI loop
    root.mainloop()

    # Return the selected path
    return selected_path


# Function to create an HTML file
def create_html(selected_path):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y %H:%M:%S")

    # Set report folder and log file paths
    report_folder = os.path.join(selected_path, "report")
    if not os.path.exists(report_folder):
        os.makedirs(report_folder)

    log_file_path = Path(report_folder + '/EBTB_TO_XOSC_Conv_Status.log')

    css_content = """


h1{
  color: rgb(12, 209, 42);
  font-family: verdana;
  font-size: 250%;
  text-align:center;
  text-underline-position:under;
}


h2{
  color: rgb(170, 13, 209);
 font-family: verdana;
 font-size: 100%;
 text-align:left;
 font-size: 22px;
 text-underline-position:under;
 font-weight: bold;
}


h3{
  color: rgb(16, 31, 168);
  font-family: verdana;
  font-size: 100%;
  text-align:left;
  font-size: 22px;
  text-underline-position:under;
  font-weight: bold;
}

h6{
  color: rgb(16, 31, 168);
  font-family: verdana;
  font-size: 100%;
  text-align:left;
  font-size: 14px;
  text-underline-position:under;
  font-weight: bold;
}

td {
  border: 0.5px solid black;
  border-collapse:collapse;
  border-style: double;
  padding: 15px;
  border: width 2px;
  text-align: center;
  text-size-adjust: 80%;
  font-weight:lighter;
  font-size: 19px;;
  color: rgb(16, 31, 168);

}


th {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(80%);
  color: rgb(12, 120, 209);
}


th.th3 {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(80%);
  color: rgb(11, 160, 18);
}


th.th8 {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  font-weight: bold;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 27px;
  filter: brightness(80%);
  color: rgb(0, 1, 8);
  background-color: rgb(21, 223, 206);
}


th.th6 {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  font-weight: bold;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 27px;
  filter: brightness(80%);
  color: rgb(0, 1, 8);
  background-color: rgb(135, 206, 235);
}

th.th7 {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  font-weight: bold;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 27px;
  filter: brightness(80%);
  color: rgb(0, 1, 8);
  background-color:rgba(255, 99, 71, 0.5)
}



th.th4 {
  border: 1px solid black;
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(80%);
  color: rgb(0, 0, 0);
}

th.th1 {
  border: 1px solid rgb(9, 65, 150);
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(80%);
  color: rgb(130, 12, 209);


}

a{
  color: rgb(16, 20, 236);
  font-family: verdana;
  text-align:right;
  text-underline-position:under;
  font-size: 27px;
  font-weight: bold;
  filter: brightness(150%);




}


th.th2 {
  border: 1px solid rgb(85, 128, 7);
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(80%);
  color: rgb(209, 12, 134);
}



td.td1 {
  border: 1px solid rgb(0, 0, 0);
  border-collapse:collapse;
  border-style: solid;
  padding: 15px;
  border: width 5px;
  text-align: center;
  text-size-adjust: 100%;
  font-size: 22px;
  filter: brightness(100%);
  color: rgb(9, 0, 15);
  font-weight: bold;
  background-color: rgb(211, 218, 219);

}


td.td2{
  border: 0.5px solid black;
  border-collapse:collapse;
  border-style: double;
  padding: 15px;
  border: width 2px;
  text-align: center;
  text-size-adjust: 80%;
  color: rgb(224, 35, 145);
  font-size: 19px;;
  background-color: rgb(226, 235, 235);
  font-weight: bold;

}


td.td3{
  border: 0.5px solid black;
  border-collapse:collapse;
  border-style: double;
  padding: 15px;
  border: width 2px;
  text-align: center;
  text-size-adjust: 80%;
  color: rgb(4, 92, 30);
  font-weight: bold;
  font-size: 24px;;
  background-color: rgb(226, 235, 235);

}


td.td4{
  border: 0.5px solid black;
  border-collapse:collapse;
  border-style: double;
  padding: 15px;
  border: width 2px;
  text-align: center;
  text-size-adjust: 80%;
  color: rgb(206, 17, 32);
  font-weight: bold;
  font-size: 24px;;
  background-color: rgb(226, 235, 235);

}

td.td5{
  border: 0.5px solid black;
  border-collapse:collapse;
  border-style: double;
  padding: 15px;
  border: width 2px;
  text-align: center;
  text-size-adjust: 80%;
  color: rgb(7, 150, 185);
  font-weight: bold;
  font-size: 24px;;
  background-color: rgb(226, 235, 235);

}

.left-align {
  text-align: left; /* Left-aligns the content in the details column */
}


x{

color: rgb(51, 204, 51)

}

y{

color: #FFA07A

}

z{

color:rgb(255, 51, 0)

}

w{

color:rgb(0, 153, 204)
}

table {

  background-color: aliceblue;

}

div.button-container {
            text-align: center; /* Center align the buttons */
            margin-top: 10px;
}

button.custom-button {
            font-size: 20px;
            padding: 10px 20px;
            margin: 10px;
            cursor: pointer;
            background-color:rgb(255, 204, 0); /* Green */
            color: black;
            border: none;
            border-radius: 8px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            transition: background-color 0.3s;
        }



"""

    with open(Path("style.css"), "w") as css_file:
        css_file.write(css_content)

    # Initialize an empty dictionary to store the structured data
    log_data_dict = {}
    current_warnings = []
    error_rows = []
    error_warnings = []
    error_description = {}

    # Open and read the log file
    if log_file_path.exists() and log_file_path.stat().st_size > 0:
        with open(log_file_path, 'r') as file:
            log_lines = file.readlines()

        # Parse the log file line by line
        for line in log_lines:

            if "ERROR - " in line:
                # Extract the part after "ERROR - "
                error_part = line.split("ERROR - ", 1)[1].strip()

                if "Error reading file" in error_part:
                    continue  # Skip this specific error

                if "[WinError 2]" in error_part:
                    continue

                if "syntax error: line 1, column 0" in error_part:
                    error_part = error_part.replace("syntax error: line 1, column 0", "APIs missing in EBTB")

                # Split into error description and file path based on the first " - "

                error_description, file_path = error_part.split(" - ", 1)
                error_description = error_description.strip()
                file_path = file_path.strip()

                file_match = file_path.split("\\")[-1]

                if file_match:
                    xosc_filename = file_match.split("\\")[-1]
                    ebtb_filename = xosc_filename.replace('.xosc', '.xebtb')

                    error_rows.append({
                        "EBTB_filename": ebtb_filename,
                        "XOSC_filename": xosc_filename,
                        "Username": getpass.getuser(),
                        "Status": "Not converted",
                        "Details": error_description
                    })
                    error_msg = f"Error - {error_description}"
                    current_warnings.append(error_msg)

                    if any("Error" in message for message in current_warnings):
                        # Filter the list to keep only the "Error" messages
                        current_warnings = [message for message in current_warnings if "Error" in message]

                    # log_data_dict[xosc_filename] = error_warnings

            elif "WARNING - " in line:
                warning_message = line.split("WARNING - ")[-1].strip()
                excluded_warnings = [
                    "No function mapped to Dri_PrepareVehicle",
                    "No function mapped to Obj_Initialize"
                ]
                if warning_message not in excluded_warnings:
                    current_warnings.append(warning_message)


            elif "INFO - Processed file in exception handler" in line:
                file_path = line.split("INFO - Processed file in exception handler: ")[-1].strip()
                file_name = file_path.split("\\")[-1]
                log_data_dict[file_name] = current_warnings
                current_warnings = []

    for key in log_data_dict.keys():
        if not log_data_dict[key]:  # If no warnings for this file
            log_data_dict[key].append("All functions mapped successfully.")

    if not log_data_dict:
        log_data_dict['No files processed'] = ['No warnings']

    # Generate repeated keys
    repeated_keys = [key for key, values in log_data_dict.items() for _ in values]
    Details = [warn for warnings in log_data_dict.values() for warn in warnings]

    status = [
        "Not converted" if "Error" in warn else
        "Partially converted" if "No function mapped" in warn else
        "Completely converted"
        for warn in Details
    ]

    # Prepare other data for HTML content
    XOSC_path = selected_path
    EBTB_paths = selected_path
    EBTB_filename = repeated_keys
    XOSC_filename = [file.replace('.xebtb', '.xosc') for file in EBTB_filename]
    Username = [getpass.getuser()] * len(Details)

    # Step 1: Group data by XOSC_filename
    grouped_data = {}

    # Iterate over the data
    for xosc_file, ebtb_file, user, stat, detail in zip(XOSC_filename, EBTB_filename, Username, status, Details):
        # Ensure EBTB_filename appears only once for each XOSC_filename

        if xosc_file not in grouped_data:
            grouped_data[xosc_file] = {
                "EBTB_filename": None,
                "Details": set(),  # Use a set to avoid duplicates
                "status": stat,
                "Username": user
            }

        if grouped_data[xosc_file]["EBTB_filename"] is None:
            grouped_data[xosc_file]["EBTB_filename"] = ebtb_file

        # Extract API functions from the detail
        api_functions = re.findall(r'No function mapped to (\S+)', detail)
        error_functions = re.findall(r'Error - (.+)', detail)
        unique_apis = set(api_functions)  # Remove duplicates

        # Add unique API functions to the set
        grouped_data[xosc_file]["Details"].update(unique_apis)
        grouped_data[xosc_file]["Details"].update(error_functions)

        # Assuming the status and username are the same for all rows with the same XOSC_filename
        grouped_data[xosc_file]["status"] = stat
        grouped_data[xosc_file]["Username"] = user

    # Get the current date and time in the desired format
    current_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    # Step 2: Prepare the merged rows for the HTML table
    merged_rows = []
    completely_converted_count = 0
    partially_converted_count = 0
    not_converted_count = 0

    for file_name, warnings in log_data_dict.items():
        status = "Completely converted"
        if any("Error" in warning for warning in warnings):
            status = "Not converted"
            not_converted_count += 1
        elif any("No function mapped" in warning for warning in warnings):
            status = "Partially converted"
            partially_converted_count += 1
        else:
            completely_converted_count += 1
    total_converted_count = partially_converted_count + completely_converted_count

    for index, (xosc_file, data) in enumerate(grouped_data.items(), start=1):

        ebtb_filenames = data["EBTB_filename"]  # EBTB_filename appears only once

        if data["status"] == "Partially converted":
            # details_list = [detail for detail in data["Details"] if detail != "Dri_PrepareVehicle"]
            # if details_list:  # Only join if there are other details
            #     details = "No function mapped to " + ", ".join(details_list)
            details = "No function mapped to " + ", ".join(
                data["Details"])  # Join all unique details for the same XOSC_filename
        elif data["status"] == "Not converted":
            error_set = data["Details"]
            details = list(error_set)[0]
            details = (f"Error - {details}")
            xosc_file = " -------------------------- "
        else:
            details = "All functions mapped for this EBTB"

        merged_rows.append(
            f"<tr><td>{index}</td><td>{ebtb_filenames}</td><td>{xosc_file}</td><td>{data['Username']}</td><td>{data['status']}</td><td class='left-align'><pre style='white-space: pre-wrap;'>{details}</pre></td></tr>")

    # The merged_rows now contains rows with unique details for each XOSC_filename
    total_count = len(merged_rows)

    # Step 3: Generate the HTML content
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <link rel="stylesheet" type="text/css" href="style.css">
        <title>EBTB to XOSC Conversion Status</title>
        <script>
        function copyToClipboard(id) {{
            const text = document.getElementById(id).innerText;
            navigator.clipboard.writeText(text).then(() => {{
                alert("Text copied to clipboard!");
            }}).catch(err => {{
                console.error("Failed to copy text: ", err);
            }}); 
        }}

        // Attempt to resize the window on load
        window.onload = function() {{
            // Set desired width and height in pixels
            const desiredWidth = 1200;
            const desiredHeight = 800;

            // Resize the window
            window.resizeTo(desiredWidth, desiredHeight);

            // Center the window on screen
            const left = (screen.width - desiredWidth) / 2;
            const top = (screen.height - desiredHeight) / 2;
            window.moveTo(left, top);
        }};
        </script>
    </head>
    <body>
        <div class="container">
            <h1>EBTB to XOSC Conversion Status</h1>
            <h2><left> <u>Generated on {current_datetime} </u></left> </h2>

                <table border="1" cellpadding="10" cellspacing="0" style="border-collapse: collapse; text-align: left;">
                <thead><tr><th>Status</th><th>Count</th></tr></thead>
                <tbody>
                <tr><td>Total EBTB Files</td><td>{total_count}</td></tr>
                <tr><td>Completely Converted </td><td>{completely_converted_count}</td></tr>
                <tr><td>Partially Converted</td><td>{partially_converted_count}</td></tr>
                <tr><td>Not Converted</td><td>{not_converted_count}</td></tr>  
                <tr><td><b>Total Converted Files</b></td><td><b>{total_converted_count}</b></td></tr>



                </tbody>
                </table>


            <h3><left> <a href = "{EBTB_paths}"> EBTB Folder </a></left> </h3>
            <h3><left> <a href = "{XOSC_path + '/report'}"> XOSC Path </a></left> </h3>

            <table>
                <thead>
                    <tr>
                        <th class="th4">S.No</th>
                        <th class="th4">EBTB_filename</th>
                        <th class="th4">XOSC_filename</th>
                        <th class="th4">Username</th>
                        <th class="th4">Status</th>
                        <th class="th4">Details</th>


                    </tr>
                </thead>
                <tbody>
                    {''.join(merged_rows)}
                </tbody>
            </table>
        </div>
    </body>
    </html>
    """

    # Step 4: Write the HTML content to the file
    html_file_path = os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.html")

    with open(os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.html"), "w") as file:
        file.write(html_content)

    # import pandas as pd
    # from openpyxl import load_workbook
    # from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    #
    # # File Paths
    # html_file_path = os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.html")
    # excel_file_path = os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.xlsx")
    #
    # # Convert HTML to DataFrame
    # dfs = pd.read_html(html_file_path)
    # if dfs:
    #     df = dfs[1]  # Assuming only one table exists
    #     df.to_excel(excel_file_path, index=False, engine="openpyxl")
    #
    #     # Load the workbook and select the active worksheet
    #     wb = load_workbook(excel_file_path)
    #     ws = wb.active
    #
    #     # Set column widths
    #     ws.column_dimensions["A"].width = 6  # "Status" column
    #     ws.column_dimensions["B"].width = 50 # "Count" column
    #     ws.column_dimensions["C"].width = 50
    #     ws.column_dimensions["D"].width = 12
    #     ws.column_dimensions["E"].width = 20
    #     ws.column_dimensions["F"].width = 35
    #
    #
    #     # Apply formatting
    #     header_font = Font(bold=True, color="FFFFFF")  # White bold text
    #     header_fill = PatternFill(start_color="9BBB59", end_color="4F81BD", fill_type="solid")  # Blue header
    #     center_align = Alignment(horizontal="left")
    #
    #     thin_border = Border(
    #         left=Side(style="thin"),
    #         right=Side(style="thin"),
    #         top=Side(style="thin"),
    #         bottom=Side(style="thin"),
    #     )
    #
    #     # Apply header formatting
    #     for cell in ws[1]:  # First row (header)
    #         cell.font = header_font
    #         cell.fill = header_fill
    #         cell.alignment = center_align
    #         cell.border = thin_border
    #
    #     # Apply formatting for data rows
    #     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    #         for cell in row:
    #             cell.alignment = center_align
    #             cell.border = thin_border
    #
    #     # Save the formatted workbook
    #     wb.save(excel_file_path)
    #
    # # Open the HTML file in the browser
    # full_path = os.path.abspath(html_file_path)
    # webbrowser.open(f"file://{full_path}")

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # File Paths
    html_file_path = os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.html")
    excel_file_path = os.path.join(report_folder, "EBTB_TO_XOSC_Conv_Status.xlsx")

    # Convert HTML to DataFrame
    dfs = pd.read_html(html_file_path)
    if dfs:
        df = dfs[1]  # Assuming only one table exists
        with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="EBTB_To_XOSC_Status")

        # Load the workbook and select the worksheet
        wb = load_workbook(excel_file_path)
        ws = wb["EBTB_To_XOSC_Status"]

        # Set column widths
        ws.column_dimensions["A"].width = 6  # "Status" column
        ws.column_dimensions["B"].width = 50  # "Count" column
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 35

        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")  # White bold text
        header_fill = PatternFill(start_color="9BBB59", end_color="4F81BD", fill_type="solid")  # Green header
        center_align = Alignment(horizontal="left")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply header formatting
        for cell in ws[1]:  # First row (header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Apply formatting for data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

        # Apply autofilter to the first row
        ws.auto_filter.ref = ws.dimensions

        # Save the formatted workbook
        wb.save(excel_file_path)

    # Open the HTML file in the browser
    full_path = os.path.abspath(html_file_path)
    webbrowser.open(f"file://{full_path}")

